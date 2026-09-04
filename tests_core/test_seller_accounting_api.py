from contextlib import contextmanager
from datetime import date
from io import BytesIO
import sqlite3

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import dependencies
from api.routers import accounting as router
from marketplace_core import seller_accounting as core
from services import accounting, db, entitlements
from tests_core.test_seller_dashboard_api import seed
from tests_core.test_postgresql_tenant_boundaries import postgres_database


def seed_operational(con):
    seed(con)
    existing={v['name'] for v in con.execute('PRAGMA table_info(accounting_order_lines)').fetchall()}
    for name in core.SAFE_COLUMNS:
        if name not in existing:
            con.execute(f'ALTER TABLE accounting_order_lines ADD COLUMN {name} TEXT')
    con.executescript('''
    ALTER TABLE accounting_manual_overrides ADD COLUMN updated_at TEXT;
    CREATE UNIQUE INDEX overrides_unique ON accounting_manual_overrides(marketplace_account_id,marketplace,row_key,field_name);
    CREATE TABLE suppliers(id INTEGER PRIMARY KEY,name TEXT);
    CREATE TABLE price_lists(id INTEGER PRIMARY KEY,owner_seller_id INTEGER,supplier_id INTEGER,local_path TEXT,source_url TEXT,last_download_at TEXT,created_at TEXT,name TEXT,active INTEGER,visibility TEXT);
    CREATE TABLE price_list_access(price_list_id INTEGER,seller_id INTEGER);
    CREATE TABLE saved_views(id INTEGER PRIMARY KEY,price_list_id INTEGER,seller_id INTEGER,snapshot_path TEXT,snapshot_storage_key TEXT,updated_at TEXT,name TEXT);
    CREATE TABLE accounting_catalog_settings(seller_id INTEGER PRIMARY KEY,configured INTEGER,updated_at TEXT);
    CREATE TABLE accounting_catalog_preferences(seller_id INTEGER,price_list_id INTEGER,enabled INTEGER,updated_at TEXT,PRIMARY KEY(seller_id,price_list_id));
    INSERT INTO suppliers VALUES(1,'Supplier');
    INSERT INTO price_lists VALUES
    (1,1,1,'own.csv','https://example.com/?secret=1','','2026-09-01','Owned',1,'private'),
    (2,2,1,'private.csv','','','2026-09-01','Private',1,'private'),
    (3,2,1,'shared.csv','','','2026-09-01','Shared',1,'shared'),
    (4,2,1,'global.csv','','','2026-09-01','Global',1,'global');
    INSERT INTO price_list_access VALUES(3,1);
    INSERT INTO saved_views VALUES(1,1,1,'view.csv','','2026-09-02','Own view');
    ''')


@pytest.fixture(params=['sqlite','postgres'])
def operational_db(request, monkeypatch):
    if request.param=='postgres':
        factory=request.getfixturevalue('postgres_database')
        from services.postgresql_backend import PostgreSQLCompatConnection
        @contextmanager
        def connect():
            with factory() as raw: yield PostgreSQLCompatConnection(raw)
    else:
        con=sqlite3.connect(':memory:',check_same_thread=False)
        con.row_factory=lambda cursor,values:dict(zip((c[0] for c in cursor.description),values))
        @contextmanager
        def connect():
            with con: yield con
    monkeypatch.setattr(db,'connect',connect)
    monkeypatch.setattr(accounting,'connect',connect)
    monkeypatch.setattr(accounting,'ensure_schema',lambda:None)
    with connect() as connection:seed_operational(connection)
    yield connect
    if request.param=='sqlite':con.close()


@pytest.fixture
def client(operational_db,monkeypatch):
    record={'id':1,'is_admin':False,'active_tenant_id':11,'tenant_ids':[11],'seller_ids':[1],'permissions':['accounting'],'tenant_role':'operator'}
    monkeypatch.setattr(dependencies,'session_user',lambda token:record)
    monkeypatch.setattr(entitlements,'feature_enabled',lambda *a:True)
    app=FastAPI();app.include_router(router.router)
    with TestClient(app) as value:yield value,record


PARAMS={'account_id':11,'date_from':'2026-09-04','date_to':'2026-09-04'}
PATH='/sellers/1/accounting'


def read(client,**filters):
    response=client.get(PATH+'/rows',params={**PARAMS,**filters})
    assert response.status_code==200,response.text
    return response.json()


def test_accounting_filters_zero_missing_and_safe_payload(client):
    http,_=client
    result=read(http,limit=1)
    accounts=http.get(PATH+'/accounts')
    assert accounts.status_code==200 and {r['id'] for r in accounts.json()}=={11,12}
    assert 'SECRET' not in accounts.text
    assert result['total']==4 and len(result['items'])==1
    assert result['totals']['sale']==230 and result['totals']['net_revenue']==50
    assert result['missing_rows']==1
    missing=read(http,missing_only=True)
    assert missing['total']==1 and missing['items'][0]['purchase_cost_eur'] is None
    cancelled=read(http,search='CANCEL')['items'][0]
    assert cancelled['sale_eur']==0 and cancelled['net_revenue_eur']==0
    assert 'PRIVATE-RAW' not in str(result) and 'SECRET' not in str(result)
    assert 'raw_json' not in result['items'][0]
    assert read(http,search='PREVIOUS')['total']==0
    assert read(http,account_id=12)['total']==1


def test_multi_filters_totals_shares_and_export_stay_on_single_seller(client,operational_db):
    http,record=client
    with operational_db() as con:
        con.execute("UPDATE accounting_order_lines SET supplier='Innpro',country_code='DE' WHERE id IN (1,2)")
        con.execute("UPDATE accounting_order_lines SET supplier='Cecotec',country_code='IT' WHERE id IN (3,4)")
        con.execute("UPDATE accounting_order_lines SET supplier='PrivateSupplier',country_code='FR' WHERE id=8")
    result=read(http,suppliers=['Innpro'],countries=['DE'],statuses=['Spedito'],limit=1)
    assert result['total']==2 and len(result['items'])==1
    assert result['totals']['sale']==150 and result['totals']['net_revenue']==50
    assert result['seller']=={'id':1,'name':'Seller A'}
    assert result['profit_split']['our_pct']==35 and result['profit_split']['partner_pct']==65
    assert result['profit_split']['our_amount']==17.5 and result['profit_split']['partner_amount']==32.5
    assert result['filter_options']['suppliers']==['Cecotec','Innpro']
    assert result['filter_options']['countries']==['DE','IT']
    assert 'PrivateSupplier' not in str(result)
    assert read(http,suppliers=['Innpro','Cecotec'])['total']==4
    assert read(http,suppliers=['Innpro'],countries=['IT'])['total']==0
    response=http.get(PATH+'/export.xlsx',params={**PARAMS,'suppliers':['Innpro'],'countries':['DE'],'statuses':['Spedito']})
    assert response.status_code==200
    book=load_workbook(BytesIO(response.content),data_only=True)
    assert sum(row[2].value=='A' for row in book.active)==2
    assert not any(row[2].value in {'CANCEL','MISSING','PRIVATE'} for row in book.active)
    book.close()
    # An authorized agency can switch clients; totals/percentages never combine them.
    record['seller_ids']=[1,2]
    record['active_tenant_type']='agency'
    second=http.get('/sellers/2/accounting/rows',params={**PARAMS,'account_id':21}).json()
    assert second['total']==1 and second['totals']['sale']==1000
    assert second['seller']=={'id':2,'name':'Seller B'}
    assert second['profit_split']['our_pct']==0 and second['profit_split']['partner_pct']==100
    assert read(http)['totals']['sale']==230


def test_edit_persists_recomputes_conflicts_and_cannot_cross_scope(client,operational_db):
    http,_=client
    item=next(r for r in read(http)['items'] if r['id']==2)
    body={'fields':{'purchase_cost_eur':30},'expected':item['edit_values']}
    response=http.patch(PATH+'/rows/2',params={'account_id':11},json=body)
    assert response.status_code==200,response.text
    row=next(r for r in read(http)['items'] if r['id']==2)
    assert row['purchase_cost_eur']==30 and row['net_revenue_eur']==15
    assert row['our_share_eur']==5.25 and row['partner_share_eur']==9.75
    assert http.patch(PATH+'/rows/2',params={'account_id':11},json=body).status_code==409
    # A later cache refresh cannot remove the authoritative manual override.
    with operational_db() as con:con.execute('UPDATE accounting_order_lines SET purchase_cost_eur=99 WHERE id=2')
    assert next(r for r in read(http)['items'] if r['id']==2)['purchase_cost_eur']==30
    for row_id,account_id in ((8,21),(8,11),(9,21)):
        assert http.patch(PATH+f'/rows/{row_id}',params={'account_id':account_id},json=body).status_code==404
    assert http.get('/sellers/2/accounting/rows',params=PARAMS).status_code==404


def test_edit_sale_recalculates_payout_and_cancel_rule_wins(client):
    http,_=client
    item=next(r for r in read(http)['items'] if r['id']==1)
    assert http.patch(PATH+'/rows/1',params={'account_id':11},json={'fields':{'sale_eur':150},'expected':item['edit_values']}).status_code==200
    item=next(r for r in read(http)['items'] if r['id']==1)
    assert item['payout_eur']==140 and item['net_revenue_eur']==80
    cancelled=read(http,search='CANCEL')['items'][0]
    response=http.patch(PATH+'/rows/3',params={'account_id':11},json={'fields':{'sale_eur':100},'expected':cancelled['edit_values']})
    assert response.status_code==200,response.text
    assert read(http,search='CANCEL')['items'][0]['sale_eur']==0


def test_catalog_union_scope_selection_and_no_secret_paths(client):
    http,_=client
    response=http.get(PATH+'/catalogs')
    assert response.status_code==200,response.text
    result=response.json()
    assert set(result['enabled_ids'])=={1,3,4} and len(result['options'])==3
    assert 'secret' not in str(result) and 'path' not in str(result) and 'source_url' not in str(result)
    assert http.put(PATH+'/catalogs',json={'enabled_ids':[2]}).status_code==404
    assert http.put(PATH+'/catalogs',json={'enabled_ids':[1]}).status_code==200
    assert http.get(PATH+'/catalogs').json()['enabled_ids']==[1]
    assert http.put(PATH+'/catalogs',json={'enabled_ids':[]}).status_code==200
    result=http.get(PATH+'/catalogs').json()
    assert result['configured'] and result['enabled_ids']==[]


def test_xlsx_all_filtered_rows_formulas_and_safe_text(client,operational_db):
    http,_=client
    with operational_db() as con:con.execute("UPDATE accounting_order_lines SET product_title='=1+1' WHERE id=1")
    response=http.get(PATH+'/export.xlsx',params=PARAMS)
    assert response.status_code==200,response.text[:200] if response.status_code!=200 else ''
    assert response.headers['cache-control']=='no-store'
    book=load_workbook(BytesIO(response.content),data_only=False)
    sheet=book.active
    assert any(cell.value=='=1+1' and cell.data_type=='s' for row in sheet for cell in row)
    # Every filtered row appears, including rows beyond the API page limit.
    assert sum(row[2].value in {'A','CANCEL','MISSING'} for row in sheet)==4
    assert any(cell.data_type=='f' for row in sheet for cell in row)
    book.close()
    assert http.get(PATH+'/export.xlsx',params={**PARAMS,'search':'UNMATCHED'}).status_code==404
    assert http.get(PATH+'/export.xlsx',params={**PARAMS,'account_id':21}).status_code==404


def test_validation_and_viewer_plan_permissions(client,monkeypatch):
    http,record=client
    item=read(http)['items'][0]
    for fields in ({'seller_id':2},{'purchase_cost_eur':'abc'},{'quantity':-1},{'quantity':1.5}):
        assert http.patch(PATH+f"/rows/{item['id']}",params={'account_id':11},json={'fields':fields,'expected':item['edit_values']}).status_code==422
    assert http.get(PATH+'/rows',params={**PARAMS,'date_to':'2026-09-01'}).status_code==422
    assert http.get(PATH+'/rows',params={**PARAMS,'limit':101}).status_code==422
    record['tenant_role']='viewer'
    assert http.put(PATH+'/catalogs',json={'enabled_ids':[]}).status_code==403
    assert http.patch(PATH+f"/rows/{item['id']}",params={'account_id':11},json={'fields':{'note':'test'},'expected':item['edit_values']}).status_code==403
    assert http.get(PATH+'/rows',params=PARAMS).status_code==200
    monkeypatch.setattr(entitlements,'feature_enabled',lambda *a:False)
    monkeypatch.setattr(entitlements,'tenant_entitlements',lambda *a:{'plan_code':'free'})
    assert http.get(PATH+'/rows',params=PARAMS).status_code==403
    record['permissions']=[]
    assert http.get(PATH+'/catalogs').status_code==403
    monkeypatch.setattr(dependencies,'session_user',lambda token:None)
    assert http.get(PATH+'/rows',params=PARAMS).status_code==401


def test_concurrent_postgresql_edits_cannot_overwrite_each_other(postgres_database,monkeypatch):
    from concurrent.futures import ThreadPoolExecutor
    from threading import Barrier
    from services.postgresql_backend import PostgreSQLCompatConnection
    @contextmanager
    def connect():
        with postgres_database() as raw:yield PostgreSQLCompatConnection(raw)
    monkeypatch.setattr(db,'connect',connect);monkeypatch.setattr(accounting,'connect',connect)
    with connect() as con:seed_operational(con)
    record=next(r for r in core.list_rows(1,11,date(2026,9,4),date(2026,9,4))['items'] if r['id']==2)
    barrier=Barrier(2)
    def save(cost):
        barrier.wait(timeout=10)
        try:
            core.save_row(1,11,2,{'purchase_cost_eur':cost},record['edit_values'])
            return 'saved'
        except accounting.AccountingEditConflict:return 'conflict'
    with ThreadPoolExecutor(max_workers=2) as pool:results=list(pool.map(save,[30,40]))
    assert sorted(results)==['conflict','saved']


def test_bulk_edit_atomic_conflict_scope_and_persistent_fields(client,operational_db):
    http,_=client
    items={r['id']:r for r in read(http)['items']}
    body={'rows':[{'id':i,'fields':{'extra_cost_eur':7,'supplier_order_number':'SUP-'+str(i),'receipt':'R'},'expected':items[i]['edit_values']} for i in (1,2)]}
    # A stale second row rolls back the already processed first row and its overrides.
    with operational_db() as con:con.execute("UPDATE accounting_order_lines SET note='Changed concurrently' WHERE id=2")
    assert http.post(PATH+'/bulk-edit',params={'account_id':11},json=body).status_code==409
    rows={r['id']:r for r in read(http)['items']}
    assert rows[1]['extra_cost_eur']==0 and rows[1]['supplier_order_number'] in ('',None)
    with operational_db() as con:
        assert con.execute('SELECT COUNT(*) AS n FROM accounting_manual_overrides').fetchone()['n']==1
    body['rows'][1]['expected']=rows[2]['edit_values']
    response=http.post(PATH+'/bulk-edit',params={'account_id':11},json=body)
    assert response.status_code==200,response.text
    assert response.json()['updated_rows']==2
    rows={r['id']:r for r in read(http)['items']}
    assert rows[1]['extra_cost_eur']==7 and rows[1]['supplier_order_number']=='SUP-1'
    assert rows[1]['net_revenue_eur']==23
    with operational_db() as con:con.execute("UPDATE accounting_order_lines SET extra_cost_eur=0,supplier_order_number='' WHERE id=1")
    assert next(r for r in read(http)['items'] if r['id']==1)['extra_cost_eur']==7
    body['rows'][1]['id']=8
    assert http.post(PATH+'/bulk-edit',params={'account_id':11},json=body).status_code==404
    assert http.post(PATH+'/bulk-edit',params={'account_id':21},json=body).status_code==404


def test_bulk_validation_permission_and_filtered_selection(client):
    http,record=client
    row=read(http)['items'][0]
    edit={'id':row['id'],'fields':{'receipt':'S'},'expected':row['edit_values']}
    for payload in ({'rows':[]},{'rows':[edit,edit]},
                    {'rows':[{**edit,'fields':{'purchase_cost_eur':3}}]},
                    {'rows':[{**edit,'fields':{'extra_cost_eur':-1}}]}):
        assert http.post(PATH+'/bulk-edit',params={'account_id':11},json=payload).status_code==422
    response=http.get(PATH+'/selection',params={**PARAMS,'search':'A'})
    assert response.status_code==200,response.text
    assert response.json()['total']==len(response.json()['items'])
    assert all(r['seller_id']==1 for r in response.json()['items'])
    assert http.get(PATH+'/selection',params={**PARAMS,'account_id':21}).status_code==404
    record['tenant_role']='viewer'
    assert http.post(PATH+'/bulk-edit',params={'account_id':11},json={'rows':[edit]}).status_code==403


def test_bulk_deleted_row_aborts_transaction(client,operational_db):
    http,_=client
    items={r['id']:r for r in read(http)['items']}
    changes=[{'marketplace_account_id':11,'marketplace':'kaufland','row_key':items[i]['row_key'],
              'fields':{'receipt':'NEW'},'expected':items[i]['edit_values']} for i in (1,2)]
    with operational_db() as con:con.execute('DELETE FROM accounting_order_lines WHERE id=2')
    with pytest.raises(accounting.AccountingEditConflict):
        accounting.save_accounting_inline_edits(changes,seller_id=1,require_all=True)
    assert next(r for r in read(http)['items'] if r['id']==1)['receipt'] in ('',None)
